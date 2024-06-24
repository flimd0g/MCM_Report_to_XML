import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import Alignment
import os

CONFIG_FILE = "config_mcm.txt"

def load_configuration():
    if os.path.isfile(CONFIG_FILE):
        with open(CONFIG_FILE, 'r') as file:
            excel_file_path = file.readline().strip()
            if excel_file_path:
                excel_file_entry.delete(0, tk.END)
                excel_file_entry.insert(0, excel_file_path)
                return excel_file_path
    return None

def save_configuration(excel_file_path):
    with open(CONFIG_FILE, 'w') as file:
        file.write(excel_file_path)

def select_excel_file():
    excel_file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
    if excel_file_path:
        excel_file_entry.delete(0, tk.END)
        excel_file_entry.insert(0, excel_file_path)
        save_configuration(excel_file_path)

def select_file():
    file_path = filedialog.askopenfilename(filetypes=[("HTML files", "*.html")])
    job_number = job_number_entry.get()
    vehicle_type = vehicle_type_var.get()
    excel_path = excel_file_entry.get()
    if file_path and job_number and vehicle_type and excel_path:
        try:
            process_file(file_path, job_number, vehicle_type, excel_path)
            messagebox.showinfo("Success", "Excel file updated successfully.")
            root.destroy()  # Close the GUI after success
        except Exception as e:
            messagebox.showerror("Error", f"An error occurred: {e}")

def parse_html(file_path):
    with open(file_path, 'r', encoding='utf-8') as file:
        soup = BeautifulSoup(file, 'html.parser')

    keys_to_extract = [
        'MCM hardware class', 'MCM version', 'MCM diagnosis version', 'MCM VIN', 'MCM serial number',
        'MCM hardware part number', 'MCM certification', 'MCM hardware version'
    ]

    extracted_values = {key: None for key in keys_to_extract}

    rows = soup.find_all('tr')
    for row in rows:
        cells = row.find_all('td')
        if len(cells) == 2:
            key = cells[0].get_text(strip=True)
            value = cells[1].get_text(strip=True)
            if key == 'MCM diagnosis version':
                value = value.lstrip('0')
            if key in extracted_values:
                extracted_values[key] = value

    for key, value in extracted_values.items():
        print(f"{key}: {value}")

    return extracted_values

def update_excel(extracted_values, job_number, vehicle_type, excel_path):
    if not os.path.isfile(excel_path):
        raise FileNotFoundError(f"Excel file not found: {excel_path}")

    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active

    header_mapping = {
        'MCM hardware class': 'Hardware Class',
        'MCM version': 'Version',
        'MCM diagnosis version': 'Diagnosis Version',
        'MCM VIN': 'Vin',
        'MCM serial number': 'Serial Number',
        'MCM hardware part number': 'Part Number',
        'MCM certification': 'Certification',
        'MCM hardware version': 'Hardware Version',
        'Job number': 'Fixably No.',
        'Vehicle Type': 'Vehicle Type'
    }

    header_row_index = None
    for row in ws.iter_rows(min_row=1, max_row=10):
        headers = {cell.value: cell.column for cell in row if cell.value}
        print(f"Headers found in row {row[0].row}: {headers}")
        if set(header_mapping.values()).issubset(headers.keys()):
            header_row_index = row[0].row
            break

    if not header_row_index:
        raise ValueError("Header row not found in the Excel sheet")

    headers = {cell.value: cell.column for cell in ws[header_row_index]}
    print(f"Headers and their columns: {headers}")

    for key in extracted_values.keys():
        if header_mapping[key] not in headers:
            raise ValueError(f"Column for '{key}' not found in the Excel sheet")

    extracted_values['Job number'] = job_number
    extracted_values['Vehicle Type'] = vehicle_type

    target_row = None
    for row in ws.iter_rows(min_row=header_row_index + 1):
        id_cell = row[0]
        if id_cell.value is not None:
            if all(cell.value is None for cell in row if cell.column != 1):
                target_row = id_cell.row
                break

    if target_row is None:
        print("No suitable row found for updating")
        raise ValueError("No suitable row found for updating")

    for key, value in extracted_values.items():
        cell = ws.cell(row=target_row, column=headers[header_mapping[key]], value=value)
        cell.alignment = Alignment(horizontal='center', vertical='center')

    wb.save(excel_path)
    wb.close()

def process_file(file_path, job_number, vehicle_type, excel_path):
    extracted_values = parse_html(file_path)
    update_excel(extracted_values, job_number, vehicle_type, excel_path)

# GUI Setup
root = tk.Tk()
root.title("HTML to Excel")

# Create a Notebook (tabbed interface)
notebook = ttk.Notebook(root)
notebook.pack(padx=10, pady=10, expand=True, fill='both')

# Main tab
main_frame = ttk.Frame(notebook)
notebook.add(main_frame, text='Main')

job_number_label = tk.Label(main_frame, text="Job number:")
job_number_label.pack()

job_number_entry = tk.Entry(main_frame)
job_number_entry.pack()

vehicle_type_label = tk.Label(main_frame, text="Vehicle Type:")
vehicle_type_label.pack()

vehicle_type_var = tk.StringVar()
vehicle_type_dropdown = ttk.Combobox(main_frame, textvariable=vehicle_type_var)
vehicle_type_dropdown['values'] = ("Truck", "Bus", "Off-Highway", "International")
vehicle_type_dropdown.pack()

select_button = tk.Button(main_frame, text="Select fault report", command=select_file)
select_button.pack()

# Configuration tab
config_frame = ttk.Frame(notebook)
notebook.add(config_frame, text='Configuration')

excel_file_label = tk.Label(config_frame, text="Excel file:")
excel_file_label.pack()

excel_file_entry = tk.Entry(config_frame)
excel_file_entry.pack()

select_excel_button = tk.Button(config_frame, text="Select Excel file", command=select_excel_file)
select_excel_button.pack()

# Load the configuration on startup
load_configuration()

root.mainloop()
